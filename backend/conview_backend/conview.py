import textwrap as tw
from pathlib import Path
from typing import NamedTuple
from collections import namedtuple, defaultdict
from dataclasses import dataclass
import tomllib

import nilearn.image as nili
import nibabel as nib
import numpy as np
import pyparadigm as pp
import pygame as pg
import matplotlib.pyplot as plt
import matplotlib as mpl
import logging
from pathlib import Path
try:
    import openpyxl  # optional, for ND.xlsx
except ImportError:  # pragma: no cover - optional dep
    openpyxl = None
# Global cache for 4D-wide value range (vmin, vmax) for the connectivity matrix,
# the threshold used to compute it, and the 4D image identity this cache belongs to.
_GLOBAL_4D_VRANGE = None
_GLOBAL_4D_THRESHOLD = None
_GLOBAL_4D_KEY = None

from .utils import (
    text, mni_2_index, index_2_mni, load_bg_template, partition_all,
    error, load_nii_or_error, threshold_slice, get_slices, cut,
    image_coords_2_mat_coords, mat_coords_2_img_coords, encode_surface)
from .core import (
    ImgCoords, MatCoords, ImgSize, MatShape, SideInfo
)
from .area_tracker import AreaTracker

menu_bar_heigth = 20
color_bar_height = 100
bar_bg_color = 0x999999
nslice = slice(None, None, None)

AtlasData = namedtuple('AtlasData', "img labels".split())
NetworkMapping = namedtuple('NetworkMapping', "order data")

@dataclass
class MatrixSummary:
    values: np.ndarray
    source_ids: list
    target_ids: list
    target_labels: list
    centers: list
    base_payload: dict

class RuntimeState:
    def __init__(self, base_path, dep_info:dict):
        self.bg_img = load_background_img()
        self.cmaps = list(plt.colormaps())

        network_cfg = dep_info.get("network_mapping") or {}
        self.region_to_network = {}

        atlas_dict = dep_info.get("atlanti")\
            or error("No section 'atlanti' in atlas-matrix-mapping")
        matrices_per_atlas = dep_info.get('compatible_matrices')\
            or error("No section 'compatible_matrices' in atlas-matrix-mapping")
        self.atlanti = {
            atlas_id: load_atlas_data(base_path / atlas_file, self.bg_img) 
            for (atlas_id, atlas_file) in atlas_dict.items() }
        self.atlas_to_matrix_map = defaultdict(dict)
        for atlas_id in self.atlanti:
            matrix_map = matrices_per_atlas.get(atlas_id)\
                or error(f'"compatible_matrices.{atlas_id}" missing from atlas-matrix-mapping')
            for matrix_id, matrix_file in matrix_map.items():
                self.atlas_to_matrix_map[atlas_id][matrix_id] = \
                    load_matrix_img(base_path / matrix_file)

        # I don't want this to be a default dict after init
        self.atlas_to_matrix_map = {**self.atlas_to_matrix_map}

        # precompute region connectivity matrices for matrix view interactions
        self.matrix_summaries = defaultdict(dict)
        for atlas_id, atlas_data in self.atlanti.items():
            for matrix_id, matrix_img in self.atlas_to_matrix_map[atlas_id].items():
                network_map = self._get_network_mapping(base_path, atlas_id, network_cfg)
                order_meta = self._get_matrix_order(base_path, atlas_id, dep_info.get("matrix_ordering") or {})
                self.matrix_summaries[atlas_id][matrix_id] = \
                    compute_region_connectivity_matrix(
                        atlas_data.img,
                        matrix_img,
                        atlas_data.labels,
                        network_map,
                        order_meta,
                    )
        self.matrix_summaries = {**self.matrix_summaries}

    def _get_network_mapping(self, base_path: Path, atlas_id: str, network_cfg: dict):
        # Prefer atlas-specific mapping from config
        path = network_cfg.get(atlas_id)
        if path:
            return load_network_mapping(base_path / path)
        # Fallback to legacy ND.xlsx / ND.csv
        nd_xlsx = base_path / "ND.xlsx"
        nd_csv = base_path / "ND.csv"
        if nd_xlsx.exists():
            return wrap_network_mapping(load_region_networks(nd_xlsx))
        if nd_csv.exists():
            return wrap_network_mapping(load_region_networks_csv(nd_csv))
        return NetworkMapping([], {})

    def _get_matrix_order(self, base_path: Path, atlas_id: str, order_cfg: dict):
        path = order_cfg.get(atlas_id)
        if path:
            p = base_path / path
            if not p.exists() and p.suffix.lower() != ".json":
                alt = p.with_suffix(".json")
                if alt.exists():
                    p = alt
            return load_matrix_order(p)
        return {}


    def get_atlas_names(self):
        return list(self.atlanti.keys())

    def get_matrix_names(self, atlas_id):
        matrix_map = self.atlas_to_matrix_map.get(atlas_id)\
            or error(f"Requested matrix map for atlas id: {atlas_id}. It doesn't exist")
        return list(matrix_map.keys())


class SessionState:
    def __init__(self):
        self.area_tracker = AreaTracker()
        self.last_renderd_index = {
            "L": None,
            "R": None
        }


class Label(NamedTuple):
    vol: str
    name: str
    

def load_labels(p: Path):
    contents = p.read_text()

    def parse_line(l):
        idx, _, value = l.partition("\t")
        elems = value.split("_")
        return (int(idx), Label(elems[1], " ".join(elems[2:])))

    return dict( parse_line(l) for l in contents.splitlines() )
    

def load_matrix_img(p):
    img = nib.load(p)
    img.get_fdata()
    return img


def load_atlas_img(img_path, bg_img):
    atlasImg = load_nii_or_error(img_path)
    atlasImg.get_fdata()
    img = nili.resample_to_img(atlasImg, bg_img, interpolation='nearest')
    img.get_fdata()
    return img


def load_background_img():
    img = load_bg_template("ch2better")
    img.get_fdata()
    return img


def compute_region_connectivity_matrix(atlas_img, matrix_img, label_map, region_network_map: NetworkMapping, order_meta: dict | None = None):
    """
    Compute region-to-region mean connectivity matrix and region centers (MNI).
    Rows correspond to source region ids (4th dimension indices), columns to atlas regions.
    """
    atlas_on_matrix = nili.resample_to_img(atlas_img, matrix_img, interpolation='nearest')
    atlas_data = atlas_on_matrix.get_fdata().astype(int)
    map_dict = region_network_map.data if region_network_map else {}
    mapping_order = region_network_map.order if region_network_map else []
    order_meta_final = dict(order_meta or {})

    network_order = order_meta_final.get("net_labels") or []
    network_order_full = order_meta_final.get("net_labels_full") or network_order

    region_ids = sorted(int(x) for x in np.unique(atlas_data) if x > 0)
    if map_dict:
        region_ids = [rid for rid in region_ids if rid in map_dict]
    if not region_ids:
        return MatrixSummary(np.zeros((0, 0)), [], [], [], [], {})

    conn = matrix_img.get_fdata()
    n_sources = conn.shape[3]
    conn_flat = conn.reshape(-1, n_sources)
    atlas_flat = atlas_data.reshape(-1)

    counts = np.bincount(atlas_flat, minlength=max(region_ids) + 1).astype(np.float64)
    counts[counts == 0] = np.nan

    sums = np.zeros((n_sources, len(counts)))
    for src_idx in range(n_sources):
        sums[src_idx] = np.bincount(
            atlas_flat, weights=conn_flat[:, src_idx], minlength=len(counts))

    counts_targets = counts[region_ids]
    with np.errstate(invalid="ignore", divide="ignore"):
        means_full = sums[:, region_ids] / counts_targets
    means_full = np.where(np.isnan(means_full), 0.0, means_full)

    centers_by_id = {}
    label_by_id = {}
    for rid in region_ids:
        lbl = None
        if rid in map_dict:
            _nd, _net, short, alt, _atlas_lbl = map_dict[rid]
            lbl = alt or short
        if lbl is None and label_map.get(rid):
            lbl = label_map.get(rid).name
        if lbl is None:
            lbl = f"Region {rid}"
        label_by_id[rid] = lbl

        coords_idx = np.argwhere(atlas_data == rid)
        if coords_idx.size == 0:
            centers_by_id[rid] = (0.0, 0.0, 0.0)
        else:
            center_idx = coords_idx.mean(axis=0)
            centers_by_id[rid] = index_2_mni(matrix_img.affine, center_idx)

    idx_by_id = {rid: idx for idx, rid in enumerate(region_ids)}
    order_pos = {rid: idx for idx, rid in enumerate(mapping_order)} if mapping_order else {}

    members_by_net = defaultdict(list)
    for rid in (mapping_order or region_ids):
        if rid in map_dict and rid in idx_by_id:
            network_name = map_dict[rid][1]
            members_by_net[network_name].append(rid)

    ordered_networks = [n for n in network_order if n in members_by_net]
    for net in members_by_net:
        if net not in ordered_networks:
            ordered_networks.append(net)

    ordered_x_ids = []
    for net in ordered_networks:
        ordered_x_ids.extend(members_by_net.get(net, []))

    for rid in region_ids:
        if rid not in ordered_x_ids:
            ordered_x_ids.append(rid)

    if not ordered_x_ids:
        ordered_x_ids = list(region_ids)

    x_idx = [idx_by_id[rid] for rid in ordered_x_ids]
    values = means_full[:, x_idx]
    region_ids = ordered_x_ids
    target_labels = [label_by_id.get(rid, f"Region {rid}") for rid in ordered_x_ids]
    centers = [centers_by_id.get(rid, (0.0, 0.0, 0.0)) for rid in ordered_x_ids]

    def src_idx_for_rid(rid: int):
        """
        Map a region id to the 4D source index. Most 4D matrices use the region
        id as the slice index (1-based in many atlases). Handle both 0- and
        1-based cases defensively and ignore out-of-range ids.
        """
        if 0 <= rid < n_sources:
            return rid
        if 0 <= (rid - 1) < n_sources:
            return rid - 1
        return None

    reversed_networks = list(reversed(ordered_networks))
    ordered_y_pairs = []
    y_counts = []
    for net in reversed_networks:
        count = 0
        for rid in members_by_net.get(net, []):
            src_idx = src_idx_for_rid(rid)
            if src_idx is None:
                continue
            ordered_y_pairs.append((rid, src_idx))
            count += 1
        y_counts.append(count)

    if not ordered_y_pairs:
        # Fallback: keep rows aligned to available sources
        for rid in ordered_x_ids:
            src_idx = src_idx_for_rid(rid)
            if src_idx is not None:
                ordered_y_pairs.append((rid, src_idx))
        y_counts = [len(ordered_y_pairs)]

    y_idx = [src for (_rid, src) in ordered_y_pairs]
    ordered_y_ids = [rid for (rid, _src) in ordered_y_pairs]

    values = values[y_idx, :]
    source_ids = ordered_y_ids
    source_slice_ids = y_idx

    boundaries_x = [0]
    for net in ordered_networks:
        boundaries_x.append(boundaries_x[-1] + len(members_by_net.get(net, [])))
    boundaries_y = [0]
    for c in y_counts:
        boundaries_y.append(boundaries_y[-1] + c)

    order_meta_final["groups"] = [members_by_net.get(net, []) for net in ordered_networks]
    order_meta_final["net_labels"] = network_order if network_order else ordered_networks
    order_meta_final["net_labels_full"] = network_order_full if network_order_full else order_meta_final["net_labels"]
    order_meta_final["net_labels_y"] = list(reversed(order_meta_final["net_labels"]))
    order_meta_final["net_labels_full_y"] = list(reversed(order_meta_final["net_labels_full"]))
    order_meta_final["net_boundaries"] = boundaries_x
    order_meta_final["net_boundaries_y"] = boundaries_y

    region_base = build_base_payload(values, source_ids, source_slice_ids, region_ids, target_labels, centers, order_meta_final, map_dict)
    network_base = build_network_payload(values, source_ids, region_ids, target_labels, centers, label_map, map_dict, order_meta_final)

    return MatrixSummary(values, source_ids, region_ids, target_labels, centers, dict(
        region=region_base,
        network=network_base
    ))


def build_base_payload(values, source_ids, source_indices, region_ids, target_labels, centers, order_meta=None, region_network_map=None):
    clean_values = [[None if np.isnan(v) else float(v) for v in row] for row in values.tolist()]
    finite_vals = values[np.isfinite(values)]
    vmin = float(finite_vals.min()) if finite_vals.size else None
    vmax = float(finite_vals.max()) if finite_vals.size else None
    target_centers_list = [list(map(float, c)) for c in centers]
    centers_by_id = dict(zip(region_ids, centers))
    x_nets = []
    y_nets = []
    source_labels = []
    source_centers = []
    for src in source_ids:
        lbl = f"Region {src}"
        if src in centers_by_id:
            source_centers.append(list(map(float, centers_by_id.get(src, (0.0, 0.0, 0.0)))))
        else:
            source_centers.append([0.0, 0.0, 0.0])
        source_labels.append(lbl)

    id_to_nd = {}
    id_to_short = {}
    id_to_alt = {}
    id_to_net = {}
    id_to_net_full = {}
    id_to_atlas = {}
    net_short_lookup = {}
    if order_meta:
        net_short = order_meta.get("net_labels") or []
        net_full = order_meta.get("net_labels_full") or []
        net_short_lookup = {full: short for full, short in zip(net_full, net_short)}
    if region_network_map:
        for rid, (nd_name, net, short, alt, atlas_lbl) in region_network_map.items():
            rid_i = int(rid)
            id_to_nd[rid_i] = nd_name
            id_to_short[rid_i] = short
            id_to_alt[rid_i] = alt
            net_full_val = net
            net_short_val = net_short_lookup.get(net_full_val, net_full_val)
            id_to_net[rid_i] = net_short_val
            id_to_net_full[rid_i] = net_full_val
            id_to_atlas[rid_i] = atlas_lbl
    x_nets = [id_to_net.get(rid) for rid in region_ids]
    y_nets = [id_to_net.get(rid) for rid in source_ids]
    x_nets_full = [id_to_net_full.get(rid) for rid in region_ids]
    y_nets_full = [id_to_net_full.get(rid) for rid in source_ids]
    centers_by_id = dict(zip(region_ids, centers))
    source_centers = [list(map(float, centers_by_id.get(src, (0.0, 0.0, 0.0)))) for src in source_ids]

    # Axis labels use alt; popups use atlas labels (fallback to alt) with network name in parens
    def popup_label(rid, fallback):
        base = id_to_atlas.get(rid, fallback)
        net = id_to_net.get(rid)
        return f"{base} ({net})" if net else base

    axis_labels = []
    popup_labels = []
    for rid, lbl in zip(region_ids, target_labels):
        short_lbl = id_to_short.get(rid) or id_to_nd.get(rid) or id_to_alt.get(rid) or lbl
        axis_labels.append(short_lbl)
        popup_base = id_to_atlas.get(rid, id_to_alt.get(rid, lbl))
        popup_labels.append(popup_base)

    popup_map = {rid: lbl for rid, lbl in zip(region_ids, popup_labels)}
    axis_map = {rid: lbl for rid, lbl in zip(region_ids, axis_labels)}

    # Axes use name_alt; popups use atlas_label (network)
    target_labels = popup_labels
    source_labels = [popup_map.get(src, id_to_alt.get(src, f"Region {src}")) for src in source_ids]
    target_short = axis_labels
    source_short = [axis_map.get(src, id_to_alt.get(src, lbl)) for src, lbl in zip(source_ids, source_labels)]

    payload = dict(
        values=clean_values,
        x_labels=target_labels,
        y_labels=source_labels,
        x_short_labels=target_short,
        y_short_labels=source_short,
        x_ids=region_ids,
        y_ids=source_indices,
        x_centers=target_centers_list,
        y_centers=source_centers,
        x_nets=x_nets,
        y_nets=y_nets,
        x_nets_full=x_nets_full,
        y_nets_full=y_nets_full,
        vmin=vmin,
        vmax=vmax,
    )
    if order_meta:
        nets = order_meta.get("net_labels")
        bounds = order_meta.get("net_boundaries")
        bounds_y = order_meta.get("net_boundaries_y")
        if nets:
            payload["net_labels"] = nets
        if order_meta.get("net_labels_full"):
            payload["net_labels_full"] = order_meta.get("net_labels_full")
        if order_meta.get("net_labels_y"):
            payload["net_labels_y"] = order_meta.get("net_labels_y")
        if order_meta.get("net_labels_full_y"):
            payload["net_labels_full_y"] = order_meta.get("net_labels_full_y")
        if bounds:
            payload["net_boundaries"] = bounds
        if bounds_y:
            payload["net_boundaries_y"] = bounds_y
        if order_meta.get("xdlim_raw"):
            payload["xdlim"] = order_meta["xdlim_raw"]
        if order_meta.get("ydlim_raw"):
            payload["ydlim"] = order_meta["ydlim_raw"]
        if order_meta.get("xlabel"):
            payload["xlabel"] = order_meta["xlabel"]
        if order_meta.get("ylabel"):
            payload["ylabel"] = order_meta["ylabel"]
    return payload


def build_network_payload(values, source_ids, region_ids, target_labels, centers, label_map, region_network_map, order_meta=None):
    if not region_network_map:
        return None
    region_id_to_network = {}
    id_to_nd = {}
    for rid, (nd_name, net, short, alt, _atlas_lbl) in region_network_map.items():
        region_id_to_network[int(rid)] = net
        id_to_nd[int(rid)] = nd_name

    if not region_id_to_network:
        return None

    net_names = sorted(set(region_id_to_network.values()))
    if not net_names:
        return None
    net_index = {n: i for i, n in enumerate(net_names)}
    net_members = [[] for _ in net_names]
    for rid, net in region_id_to_network.items():
        if net in net_index:
            net_members[net_index[net]].append(int(rid))

    sums = np.zeros((len(net_names), len(net_names)))
    counts = np.zeros_like(sums)

    for si, src in enumerate(source_ids):
        net_s = region_id_to_network.get(src)
        if net_s is None:
            continue
        s_idx = net_index[net_s]
        for ti, tgt in enumerate(region_ids):
            net_t = region_id_to_network.get(tgt)
            if net_t is None:
                continue
            val = values[si, ti]
            if np.isnan(val):
                continue
            t_idx = net_index[net_t]
            sums[s_idx, t_idx] += val
            counts[s_idx, t_idx] += 1

    with np.errstate(invalid="ignore", divide="ignore"):
        means = sums / counts
    means[counts == 0] = np.nan

    net_centers = []
    centers_by_region = dict(zip(region_ids, centers))
    for net in net_names:
        member_centers = [centers_by_region[rid] for rid, n in region_id_to_network.items() if n == net and rid in centers_by_region]
        if not member_centers:
            net_centers.append([0.0, 0.0, 0.0])
        else:
            net_centers.append(list(map(float, np.mean(member_centers, axis=0))))

    net_labels = None
    net_labels_full = None
    if order_meta:
        net_labels = order_meta.get("net_labels")
        net_labels_full = order_meta.get("net_labels_full")

    display_axis = net_labels if net_labels and len(net_labels) == len(net_names) else net_names
    display_full = net_labels_full if net_labels_full and len(net_labels_full) == len(net_names) else display_axis

    net_short = display_axis

    region_idx_map = {rid: idx for idx, rid in enumerate(region_ids)}
    net_member_indices = []
    for members in net_members:
        net_member_indices.append([region_idx_map.get(r, -1) for r in members if region_idx_map.get(r, -1) is not None and region_idx_map.get(r, -1) >= 0])

    # Flip rows (Y) to mirror region orientation (x normal, y reversed)
    rev_idx = list(range(len(net_names) - 1, -1, -1))
    means = means[rev_idx, :]
    y_labels = [display_axis[i] for i in rev_idx]
    y_short = [net_short[i] for i in rev_idx]
    y_ids = [list(range(len(net_names)))[i] for i in rev_idx]
    y_centers = [net_centers[i] for i in rev_idx]

    clean_values = [[None if np.isnan(v) else float(v) for v in row] for row in means.tolist()]

    return dict(
        values=clean_values,
        x_labels=display_axis,
        y_labels=y_labels,
        x_short_labels=net_short,
        y_short_labels=y_short,
        x_ids=list(range(len(net_names))),
        y_ids=y_ids,
        x_centers=net_centers,
        y_centers=y_centers,
        net_members=net_members,
        net_member_indices=net_member_indices,
        net_labels=display_axis,
        net_labels_full=display_full,
        net_labels_y=y_labels,
        net_labels_full_y=display_full[::-1],
        vmin=float(np.nanmin(means)) if np.isfinite(means).any() else None,
        vmax=float(np.nanmax(means)) if np.isfinite(means).any() else None,
    )


def load_region_networks(path: Path):
    """
    Load optional ND.xlsx mapping Region label -> Network.
    Expected columns: 'Region label' and 'Network'
    """
    if not path.exists():
        logging.info("No ND.xlsx found at %s; skipping network aggregation", path)
        return {}
    if openpyxl is None:
        logging.warning("openpyxl not installed; cannot read ND.xlsx")
        return {}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        col_map = {name.lower(): idx for idx, name in enumerate(headers)}
        required = ["region label", "network"]
        if not all(r in col_map for r in required):
            logging.warning("ND.xlsx missing required headers %s", required)
            return {}
        res = {}
        for row in sheet.iter_rows(min_row=2):
            region_label = row[col_map["region label"]].value
            network = row[col_map["network"]].value
            if region_label and network:
                res[str(region_label).strip()] = str(network).strip()
        return res
    except Exception as e:  # pragma: no cover
        logging.warning("Failed to read ND.xlsx: %s", e)
        return {}


def load_region_networks_csv(path: Path):
    """
    Load ND.csv or custom CSV mapping. Expected headers: Region, Network
    """
    import csv
    if not path.exists():
        return {}
    res = {}
    with path.open() as f:
        reader = csv.DictReader(f)
        for row in reader:
            region = (row.get("Region") or row.get("region") or "").strip()
            net = (row.get("Network") or row.get("network") or "").strip()
            if region and net:
                res[region] = net
    return res


def wrap_network_mapping(data: dict) -> NetworkMapping:
    order = list(data.keys())
    return NetworkMapping(order, data)


def load_network_mapping(path: Path):
    """
    Load a custom CSV mapping with columns: id, nd_region, network, shortname, name_alt, atlas_label
    """
    import csv
    if not path.exists():
        logging.info("Network mapping %s not found", path)
        return {}
    res = {}
    order = []
    with path.open() as f:
        reader = csv.DictReader(f)
        raw_rows = []
        for row in reader:
            try:
                rid = int(row.get("id") or 0)
            except Exception:
                continue
            raw_rows.append((rid, row))

        # Detect indexing: shift by +1 only if data appears zero-based
        need_shift = any(rid == 0 for rid, _ in raw_rows)

        for rid, row in raw_rows:
            rid_use = rid + 1 if need_shift else rid
            nd_region = (row.get("nd_region") or row.get("Region") or row.get("region") or "").strip()
            net = (row.get("network") or row.get("Network") or "").strip()
            shortname = (row.get("shortname") or "").strip()
            name_alt = (row.get("name_alt") or "").strip()
            atlas_label = (row.get("atlas_label") or "").strip()
            if not net or net.lower() in ("none", "nan", "na", "n/a", "missing"):
                continue
            if rid_use and net:
                res[rid_use] = (
                    nd_region,
                    net,
                    shortname if shortname else nd_region,
                    name_alt if name_alt else nd_region,
                    atlas_label if atlas_label else nd_region
                )
                order.append(rid_use)
    return NetworkMapping(order, res)


def load_matrix_order(path: Path):
    if not path.exists():
        logging.info("Matrix order file %s not found", path)
        return {}
    try:
        groups = []
        if path.suffix.lower() == ".json":
            import json
            data = json.loads(path.read_text())
        else:
            data = tomllib.loads(path.read_text())

        order_raw = data.get("order")
        if data.get("final_index"):
            groups = []
            for grp in data.get("final_index"):
                groups.append([int(i) for i in grp])
            if order_raw is None:
                order_raw = [i for grp in groups for i in grp]

        net_labels = data.get("net_labels") or []
        net_labels_full = data.get("net_labels_full") or []

        # Derive net boundaries from groups if available
        net_bounds = []
        if groups:
            cum = 0
            net_bounds = [0]
            for grp in groups:
                cum += len(grp)
                net_bounds.append(cum)
        else:
            net_bounds = data.get("net_boundaries") or []

        xdlim = data.get("xdlim") or []
        ydlim = data.get("ydlim") or []
        xlabel = data.get("xlabel") or []
        ylabel = data.get("ylabel") or []
        inv_groups = data.get("final_index_inverse")
        if inv_groups:
            inv_groups = [[int(i) for i in grp] for grp in inv_groups]

        return {
            "order": order_raw or [],
            "groups": groups,
            "final_index_inverse": inv_groups,
            "net_labels": net_labels,
            "net_labels_full": net_labels_full,
            "net_boundaries": net_bounds,
            "xdlim_raw": xdlim,
            "ydlim_raw": ydlim,
            "xlabel": xlabel,
            "ylabel": ylabel,
        }
    except Exception as e:  # pragma: no cover
        logging.warning("Failed to read matrix order %s: %s", path, e)
        return {}


def short_label(name: str, max_len: int = 8) -> str:
    """
    Make a compact label using the first two letters of each word/segment.
    Preserves L/R tokens (makes them upper-case) for later highlighting.
    """
    if not name:
        return ""
    tokens = [t for t in name.replace("_", " ").split(" ") if t]
    pieces = []
    for t in tokens:
        token = t.upper() if t.lower() in ("l", "r", "left", "right") else t
        pieces.append(token[:2])
    short = " ".join(pieces) if pieces else name[:max_len]
    return short[:max_len]


"""def render_img(atlas_img: nib.Nifti1Image, bg_img:nib.Nifti1Image, four_d_img: nib.Nifti1Image, idx_4d: int, 
               infos_l: SideInfo, infos_r: SideInfo, surface: pg.Surface, session_state: SessionState, label_map):
    atlas_mat, atlas_normalizer = apply_thresholding_and_cmap(atlas_img.get_fdata(), infos_l)
    connectivities = get_3d_slice(four_d_img, idx_4d, bg_img)
    conn_mat, conn_normalizer = apply_thresholding_and_cmap(connectivities, infos_r)
    l_idx = mni_2_index(bg_img.affine, infos_l.coords)
    value_l = atlas_img.get_fdata()[l_idx]
    r_idx = mni_2_index(bg_img.affine, infos_r.coords)
    value_r = connectivities[r_idx]
    label_l = label_map[int(value_l)] if value_l > 0 else None
    region_index_r = int(atlas_img.get_fdata()[r_idx])
    label_r = label_map[region_index_r] if region_index_r > 0 else None
    selected_col_id = region_index_r if region_index_r > 0 else None
    
    surface.fill(0xFFFFFF)
    pp.compose(surface, pp.FreeFloatLayout())(
        pp.FRect(0, 0, 0.5, -color_bar_height)(
            pp.Overlay(
                render_coord_info(infos_l.coords, int(value_l)),
                render_view(atlas_mat, bg_img, l_idx, infos_l.cmap, infos_l.smoothed, session_state, "L"))),

        pp.FRect(0, -color_bar_height, 0.5, color_bar_height)(
            make_scale(atlas_normalizer, infos_l.cmap)),

        pp.FRect(0.5, 0, 0.5, -color_bar_height)(
            pp.Overlay(
                render_coord_info(infos_r.coords, f"{value_r:.3f}"),
                render_view(conn_mat, bg_img, r_idx, infos_r.cmap, infos_r.smoothed, session_state, "R"))),

        pp.FRect(0.5, -color_bar_height, 0.5, color_bar_height)(
            make_scale(conn_normalizer, infos_r.cmap)))
    return label_l, label_r, connectivities, r_idx, selected_col_id, value_l, value_r
"""
def _four_d_cache_key(four_d_img: nib.Nifti1Image):
    """Best-effort identity for a 4D image to decide when to recompute global ranges."""
    fname = None
    try:
        fname = four_d_img.get_filename()
    except Exception:
        fname = None
    if fname:
        return ("file", str(Path(fname).resolve()))
    return ("id", id(four_d_img))

def resolve_connectivity_vrange(infos_r: SideInfo, four_d_img: nib.Nifti1Image) -> SideInfo:
    """
    For the connectivity (right) panel, upgrade an auto vrange to a global 4D
    range and return a new SideInfo (SideInfo is immutable).
    """
    global _GLOBAL_4D_VRANGE, _GLOBAL_4D_THRESHOLD, _GLOBAL_4D_KEY

    vmin, vmax = infos_r.vrange
    auto_requested = vmin is None or vmax is None or (vmin, vmax) == (-2, 2)
    if not auto_requested:
        return infos_r

    def _normalize_threshold(thresh):
        if thresh is None:
            return (None, None)
        if isinstance(thresh, (list, tuple)) and len(thresh) == 2:
            low, high = thresh
            return (low, high)
        return (None, None)

    threshold = _normalize_threshold(infos_r.threshold)
    key = _four_d_cache_key(four_d_img)
    need_recompute = (
        _GLOBAL_4D_VRANGE is None
        or _normalize_threshold(_GLOBAL_4D_THRESHOLD) != threshold
        or _GLOBAL_4D_KEY != key
    )
    if need_recompute:
        try:
            data4d = four_d_img.get_fdata()
            _GLOBAL_4D_VRANGE = compute_global_vrange(data4d, threshold)
            _GLOBAL_4D_THRESHOLD = threshold
            _GLOBAL_4D_KEY = key
            print(f"[CMAP] 4D global vrange cached: vmin={_GLOBAL_4D_VRANGE[0]:.4f}, vmax={_GLOBAL_4D_VRANGE[1]:.4f}")
        except Exception as e:  # pragma: no cover - defensive guard
            print(f"[CMAP] Failed to compute 4D global vrange: {e}")
            return infos_r

    vmin4d, vmax4d = _GLOBAL_4D_VRANGE
    return infos_r._replace(vrange=(vmin4d, vmax4d))

def render_img(atlas_img: nib.Nifti1Image, bg_img:nib.Nifti1Image, four_d_img: nib.Nifti1Image, idx_4d: int, 
               infos_l: SideInfo, infos_r: SideInfo, surface: pg.Surface, session_state: SessionState, label_map):
    # --- LEFT: atlas panel (unchanged behavior: auto per-atlas) ---
    atlas_mat, atlas_normalizer = apply_thresholding_and_cmap(atlas_img.get_fdata(), infos_l)

    # --- RIGHT: 4D relevance matrix panel ---
    # If vrange is "auto" (None/None or (-2, 2)), compute a TRUE global range
    # from the full 4D matrix and pin it as explicit vrange for the right panel.
    infos_r = resolve_connectivity_vrange(infos_r, four_d_img)

    connectivities = get_3d_slice(four_d_img, idx_4d, bg_img)
    conn_mat, conn_normalizer = apply_thresholding_and_cmap(connectivities, infos_r)

    l_idx = mni_2_index(bg_img.affine, infos_l.coords)
    value_l = atlas_img.get_fdata()[l_idx]
    r_idx = mni_2_index(bg_img.affine, infos_r.coords)
    value_r = connectivities[r_idx]
    label_l = label_map[int(value_l)] if value_l > 0 else None
    region_index_r = int(atlas_img.get_fdata()[r_idx])
    label_r = label_map[region_index_r] if region_index_r > 0 else None
    selected_col_id = region_index_r if region_index_r > 0 else None
    color_l = value_to_hex(value_l, atlas_normalizer, infos_l.cmap)
    color_r = value_to_hex(value_r, conn_normalizer, infos_r.cmap)
    vrange_l = (float(atlas_normalizer.vmin), float(atlas_normalizer.vmax))
    vrange_r = (float(conn_normalizer.vmin), float(conn_normalizer.vmax))
    
    surface.fill(0xFFFFFF)
    pp.compose(surface, pp.FreeFloatLayout())(
        pp.FRect(0, 0, 0.5, -color_bar_height)(
            pp.Overlay(
                render_coord_info(infos_l.coords, int(value_l)),
                render_view(atlas_mat, bg_img, l_idx, infos_l.cmap, infos_l.smoothed, session_state, "L"))),

        pp.FRect(0, -color_bar_height, 0.5, color_bar_height)(
            make_scale(atlas_normalizer, infos_l.cmap)),

        pp.FRect(0.5, 0, 0.5, -color_bar_height)(
            pp.Overlay(
                render_coord_info(infos_r.coords, f"{value_r:.3f}"),
                render_view(conn_mat, bg_img, r_idx, infos_r.cmap, infos_r.smoothed, session_state, "R"))),

        pp.FRect(0.5, -color_bar_height, 0.5, color_bar_height)(
            make_scale(conn_normalizer, infos_r.cmap)))
    return label_l, label_r, connectivities, r_idx, selected_col_id, value_l, value_r, color_l, color_r, vrange_l, vrange_r

def compute_global_vrange(data: np.ndarray, threshold, default=(-1.0, 1.0)):
    """
    Compute global (vmin, vmax) over an entire array after thresholding.

    Parameters
    ----------
    data : np.ndarray
        3D or 4D array of relevance/importance values.
    threshold : (low, high) or None
        Same format as in threshold_slice.
    default : (float, float)
        Fallback if everything is NaN / empty after thresholding.

    Returns
    -------
    (vmin, vmax) : tuple[float, float]
    """
    arr = np.asarray(data, dtype=float)

    # Apply the same threshold rule globally
    if threshold is not None:
        arr = threshold_slice(arr, threshold)

    finite = arr[np.isfinite(arr)]
    if finite.size:
        gmin = float(finite.min())
        gmax = float(finite.max())
        if gmin == gmax:
            span = abs(gmin) if gmin != 0 else 1.0
            gmin -= 0.1 * span
            gmax += 0.1 * span
        return gmin, gmax

    return default


def value_to_hex(val, normalizer: mpl.colors.Normalize, cmap_name: str):
    """
    Map a scalar value to a hex color using the provided normalizer + cmap.
    Returns None if mapping fails.
    """
    try:
        if val is None or np.isnan(val):
            return None
        normed = float(normalizer(val))
        cmap = mpl.cm.get_cmap(cmap_name)
        r, g, b, _ = cmap(normed)
        return "#{:02x}{:02x}{:02x}".format(
            int(round(r * 255)),
            int(round(g * 255)),
            int(round(b * 255)),
        )
    except Exception as e:  # pragma: no cover - defensive
        print(f"[CMAP] Failed to map value to color for {cmap_name}: {e}")
        return None
def apply_thresholding_and_cmap(mat3d, infos: SideInfo):
    """
    Apply thresholding + colormap normalization.

    Behavior:
      - If infos.vrange == (-2, 2) and a global 4D range is cached, use that
        (true 4D-global scaling for the connectivity panel).
      - Else if infos.vrange is auto ((None, None) or (-2, 2) without global),
        compute a local range from this slice.
      - Else (explicit vrange): respect it and only guard against zero span.

    Also prints which mode was used.
    """
    global _GLOBAL_4D_VRANGE

    data = np.asarray(mat3d, dtype=float)

    # Threshold this volume/slice
    data_thr = threshold_slice(data, infos.threshold)

    vmin, vmax = infos.vrange
    is_auto = (vmin is None or vmax is None or (vmin, vmax) == (-2, 2))

    # --------------------------------------------------------
    # 4D GLOBAL MODE (connectivity pane with sentinel (-2, 2))
    # --------------------------------------------------------
    if (vmin, vmax) == (-2, 2) and _GLOBAL_4D_VRANGE is not None:
        vmin, vmax = _GLOBAL_4D_VRANGE
        print(f"[CMAP] Using 4D GLOBAL range: vmin={vmin:.4f}, vmax={vmax:.4f}")

    # --------------------------------------------------------
    # AUTO MODE but no 4D global cached → per-slice range
    # (this will be used e.g. by the atlas with vrange = (None, None),
    #  and also by connectivity if global computation somehow failed)
    # --------------------------------------------------------
    elif is_auto:
        finite = data_thr[np.isfinite(data_thr)]
        if finite.size:
            auto_min = float(np.min(finite))
            auto_max = float(np.max(finite))
            vmin, vmax = auto_min, auto_max
        else:
            vmin, vmax = -1.0, 1.0
        print(f"[CMAP] Using LOCAL auto range: vmin={vmin:.4f}, vmax={vmax:.4f}")

    # --------------------------------------------------------
    # EXPLICIT vrange (non-auto) → respect it
    # --------------------------------------------------------
    else:
        print(f"[CMAP] Using EXPLICIT range: vmin={vmin:.4f}, vmax={vmax:.4f}")

    # Guard against zero span
    if vmin == vmax:
        span = abs(vmin) if vmin != 0 else 1.0
        vmin -= 0.1 * span
        vmax += 0.1 * span

    normalizer = mpl.colors.Normalize(vmin, vmax)
    masked_mat = np.ma.masked_invalid(data_thr)
    mat = normalizer(masked_mat)
    return mat.filled(np.nan), normalizer



"""def apply_thresholding_and_cmap(mat3d, infos: SideInfo):
    mat3d = threshold_slice(mat3d, infos.threshold)
    vmin, vmax = infos.vrange
    finite = mat3d[np.isfinite(mat3d)]
    if finite.size:
        auto_min = float(np.min(finite))
        auto_max = float(np.max(finite))
        if vmin is None or vmax is None or (vmin, vmax) == (-2, 2):
            vmin, vmax = auto_min, auto_max
        if vmin == vmax:
            span = abs(vmin) if vmin != 0 else 1.0
            vmin -= 0.1 * span
            vmax += 0.1 * span
    else:
        if vmin is None:
            vmin = -1.0
        if vmax is None:
            vmax = 1.0
        if vmin == vmax:
            vmin -= 1.0
            vmax += 1.0
    normalizer = mpl.colors.Normalize(vmin, vmax)
    masked_mat = np.ma.masked_invalid(mat3d)
    mat = normalizer(masked_mat)
    return (mat.filled(np.nan), normalizer)
"""
    
"""def get_3d_slice(img4d, idx, bg_img):
    if not 0 <= idx < img4d.shape[3]:
        raise ValueError("Invalid 4D index")
    mat = img4d.dataobj[:, :, :, idx]
    img = nib.Nifti1Image(mat, img4d.affine)
    img = nili.resample_to_img(
        img, bg_img, interpolation='nearest')
    return img.get_fdata()"""
def get_3d_slice(img4d, idx, bg_img):
    """
    Extract a 3D volume from a 4D NIfTI and resample to bg_img space.

    On the first call, also compute the global (vmin, vmax) over the entire
    4D image and cache it in _GLOBAL_4D_VRANGE so the connectivity color
    scale can be kept stable across slices.
    """
    global _GLOBAL_4D_VRANGE, _GLOBAL_4D_THRESHOLD, _GLOBAL_4D_KEY

    if not 0 <= idx < img4d.shape[3]:
        raise ValueError("Invalid 4D index")

    # Reset global cache if a different 4D image is being used
    key = _four_d_cache_key(img4d)
    if key != _GLOBAL_4D_KEY:
        _GLOBAL_4D_VRANGE = None
        _GLOBAL_4D_THRESHOLD = None
        _GLOBAL_4D_KEY = key

    # Compute and cache global 4D range once
    if _GLOBAL_4D_VRANGE is None:
        try:
            data4d = img4d.get_fdata()
            # If you want threshold-aware global, pass infos_r.threshold instead of None
            vmin4d, vmax4d = compute_global_vrange(data4d, threshold=None)
            _GLOBAL_4D_VRANGE = (vmin4d, vmax4d)
            print(f"[CMAP] 4D global range computed: vmin={vmin4d:.4f}, vmax={vmax4d:.4f}")
        except Exception as e:
            print(f"[CMAP] Failed to compute 4D global range: {e}")
            _GLOBAL_4D_VRANGE = None

    # Extract and resample the requested 3D volume
    mat = img4d.dataobj[:, :, :, idx]
    img = nib.Nifti1Image(mat, img4d.affine)
    img = nili.resample_to_img(img, bg_img, interpolation='nearest')
    return img.get_fdata()


def render_view(fg_mat, bg_img, index_coords, cmap,  
                smooth, session_state, area_prefix):
    session_state.last_renderd_index[area_prefix] = index_coords
    fg_splits = get_slices(fg_mat, index_coords)
    bg_splits = get_slices(bg_img.get_fdata(), index_coords)
    imgs = [make_overlayed_img(fg, bg, cmap, smooth)
        for fg, bg in zip(fg_splits, bg_splits)]

    for i, img in enumerate(imgs):
        imgCoords = compute_image_cross_coord(
            index_coords, i, bg_img.shape, ImgSize.from_(img))
        add_split_cross(img, imgCoords, target=img)

    return compose_3_way_plit_22(imgs, session_state.area_tracker, area_prefix)


def compute_image_cross_coord(index_coords, slice_dimension, nii_shape, imgShape):
    matCoords = MatCoords(*cut(index_coords, slice_dimension))
    matShape = MatShape(*cut(nii_shape, slice_dimension))
    return mat_coords_2_img_coords(matCoords, matShape, imgShape)


def add_split_cross(img: pg.Surface, cross_coord: ImgCoords, 
                    color: pg.Color = 0, width: int = 2, target=None):
    target = target or img.copy()
    pg.draw.line(target, color, 
                 (cross_coord.x, 0), 
                 (cross_coord.x, img.get_height()),
                 width)
    pg.draw.line(target, color, 
                 (0, cross_coord.y), 
                 (img.get_width(), cross_coord.y),
                 width)
    return target


def compose_3_way_plit_22(imgs, area_tracker, prefix):
    assert len(imgs) == 3
    for i, img in enumerate(imgs):
        imgs[i] = pp.make_transparent_by_colorkey(img, img.get_at((0, 0)))

    def field(index):
        return area_tracker.area(f"{prefix}-{index}")(pp.Surface(scale=1)(imgs[index]))

    return pp.GridLayout()(
        # the order is due to conventions in the Neuroimaging field
        [field(1), field(0)],
        [field(2), None]
    )


def make_overlayed_img(fgMat, bgMat, cmap, smooth):
    assert len(fgMat.shape) == len(bgMat.shape) == 2
    mask = np.isnan(fgMat)
    safe_fg = np.where(mask, 0.0, fgMat)
    img = pp.mat_to_surface(safe_fg, pp.apply_color_map(cmap, normalize=False)).convert_alpha()
    img = pp.make_transparent_by_colorkey(img, img.get_at((0, 0)))
    img = pg.transform.rotate(img, 90)
    mask_for_surface = mask.T  # after rotation, surface width/height match transposed mask

    bg = pp.mat_to_surface(bgMat)
    bg = pp.make_transparent_by_colorkey(bg, bg.get_at((0, 0)))
    bg = pg.transform.rotate(bg, 90)

    fg_rgb = pg.surfarray.pixels3d(img)
    fg_alpha = pg.surfarray.pixels_alpha(img)
    bg_rgb = pg.surfarray.pixels3d(bg)
    if mask_for_surface.shape == fg_rgb.shape[:2]:
        fg_rgb[mask_for_surface] = bg_rgb[mask_for_surface]
        fg_alpha[mask_for_surface] = 255
    del fg_rgb, fg_alpha, bg_rgb
    result_size = (max(img.get_width(), bg.get_width()), 
                   max(img.get_height(), bg.get_height()))
    return pp.compose(pp.empty_surface(0xFFFFFF, result_size))(
        pp.Overlay(
            pp.Surface(smooth=smooth, scale=1)(bg),
            pp.Surface(smooth=smooth, scale=1)(img),
        ))


def make_color_map_surface(normalizer, color_map):
    """
    Build a colorbar surface that reflects the current normalization range,
    not a hard-coded [-1, 1] span.
    """
    vmin = normalizer.vmin if normalizer.vmin is not None else -1.0
    vmax = normalizer.vmax if normalizer.vmax is not None else 1.0
    vals = np.linspace(vmin, vmax, 256)
    normed = normalizer(vals)
    surf = pp.mat_to_surface(
        np.expand_dims(normed, axis=0),
        pp.apply_color_map(color_map, normalize=False),
    )
    return pp.Surface(scale=1, keep_aspect_ratio=False)(surf)


def make_scale(normalizer, color_map):
    return pp.LinLayout("v")(
            pp.LLItem(2),
            pp.LLItem(2)(pp.Padding(0.1, 0.1, 0, 0)(make_color_map_surface(normalizer, color_map))),
            pp.LLItem(1),
            pp.LLItem(4)(
                pp.Padding(0.1, 0.1, 0, 0)(
                    pp.LinLayout("h")(
                        pp.LLItem(0.1),
                        text(f"{normalizer.vmin:7.2f}", align="center"),
                        pp.LLItem(2),
                        text(f"{0.5 * (normalizer.vmin + normalizer.vmax):7.2f}", align="center"),
                        pp.LLItem(2),
                        text(f"{normalizer.vmax:7.2f}", align="center")))))


def render_coord_info(mni_coords, val):
    return pp.Padding.from_scale(0.7)(
        text("")
    )

def load_atlas_data(atlas_img_path, bg_img):
    atlas_map = load_atlas_img(atlas_img_path, bg_img)
    labels = load_labels(atlas_img_path.with_suffix(".txt"))
    return AtlasData(atlas_map, labels)
